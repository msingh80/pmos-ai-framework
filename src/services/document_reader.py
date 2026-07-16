from pypdf import PdfReader
from openpyxl import load_workbook


class DocumentReader:

    @staticmethod
    def read_documents(documents):

        content = ""

        for document in documents:

            file = document["path"]

            try:

                content += (
                    f"\n\n===== "
                    f"{document['name']} "
                    f"=====\n\n"
                )

                if file.endswith(".md"):

                    with open(
                        file,
                        "r",
                        encoding="utf-8"
                    ) as f:

                        content += f.read()

                elif file.endswith(".pdf"):

                    pdf = PdfReader(file)

                    for page in pdf.pages:

                        text = page.extract_text()

                        if text:

                            content += (
                                text + "\n"
                            )

                elif file.endswith(".xlsx"):

                    workbook = load_workbook(
                        file,
                        data_only=True
                    )

                    for sheet in workbook:

                        content += (
                            f"\n--- SHEET: "
                            f"{sheet.title} ---\n"
                        )

                        for row in sheet.iter_rows(
                            values_only=True
                        ):

                            content += (
                                str(row) + "\n"
                            )

                else:

                    content += (
                        "\nSkipped file type.\n"
                    )

            except Exception as error:

                content += (
                    f"\nERROR READING FILE: "
                    f"{error}\n"
                )

        return content