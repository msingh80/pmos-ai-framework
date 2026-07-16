from pathlib import Path

from pypdf import PdfReader
from openpyxl import load_workbook


class DocumentReader:

    @staticmethod
    def read_documents():

        root = Path(__file__).resolve().parents[2]

        input_folder = root / "INPUT_DOCUMENTS"

        content = ""

        if not input_folder.exists():

            return content

        for file in input_folder.rglob("*"):

            if file.is_file():

                content += (
                    f"\n\n===== {file.relative_to(input_folder)} =====\n\n"
                )

                try:

                    if file.suffix == ".md":

                        with open(
                            file,
                            "r",
                            encoding="utf-8"
                        ) as f:

                            content += f.read()

                    elif file.suffix == ".pdf":

                        pdf = PdfReader(file)

                        for page in pdf.pages:

                            text = page.extract_text()

                            if text:

                                content += (
                                    text + "\n"
                                )

                    elif file.suffix == ".xlsx":

                        workbook = load_workbook(
                            file,
                            data_only=True
                        )

                        for sheet in workbook:

                            content += (
                                f"\n--- SHEET: "
                                f"{sheet.title} ---\n"
                            )

                            for row in sheet.iter_rows(
                                values_only=True
                            ):

                                content += (
                                    str(row) + "\n"
                                )

                    else:

                        content += (
                            "\nSkipped file type.\n"
                        )

                except Exception as error:

                    content += (
                        f"\nERROR READING FILE: "
                        f"{error}\n"
                    )

        return content